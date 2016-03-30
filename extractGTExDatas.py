from xlrd import *
from xlwt import *
from xlutils.copy import *
import time
import numpy as np
import csv

import xlsxwriter as xlwriter

import openpyxl

import paths


def extractGene(fileName):
    '''Extract SAMPID and GeneID'''
    geneNameList = []
    geneDescriptionList = []
    sampleList = []
    i = 0
    file = open(fileName,'r')
    for line in file:
        print i
        element = ''
        j = 0
        for c in line:
            element += c
            if i == 2:
                if c == '	':
                    sampleList.append(element)
                    element = ''
            else:
                if c == '	':
                    if j==0:
                        geneNameList.append(element)
                    else:
                        geneDescriptionList.append(element)
                        break
                    element = ''
                    j+=1
        i += 1
    return sampleList, geneNameList, geneDescriptionList

def extractTranscript(fileName):
    '''extract TranscriptID and chromosome and coord of Gene'''
    transcriptIDList = []
    geneSymbolList = []
    chromosomesList = []
    coordList = []

    i = 0

    file = open(fileName, 'r')
    for line in file:
        print i
        element = ''
        j = 0
        for c in line:
            element+=c
            if c == '	':
                if j == 0:
                    transcriptIDList.append(element)
                if j == 1:
                    geneSymbolList.append (element)
                elif j == 2:
                    chromosomesList.append (element)
                elif j == 3:
                    coordList.append(element)
                    break
                element =''
                j+=1
        i+=1
    return transcriptIDList, geneSymbolList, chromosomesList, coordList

def extractRPKMGene():
    return

def cleanList(list, listType):
    '''delete elements of list which not correspond to the type of the list'''
    comparison = False
    for element in list:
        if listType == 'gene':
            comparison = element[0:4] != 'GTEX'
        elif listType == 'sample':
            comparison = element[0:4] != 'ENSG'
        if comparison: #pourquoi ne connait pas l'element Description???? (element juste avant ceux a garder
            list.remove(element)
    return

def writeSampleTable():
    '''write table (excel) containing sample informations'''
    sampleTable = open_workbook(sampleTableFileName)
    samplePhenotypeTable = open_workbook(samplePhenotypeTableFileName)

    sampleTableCopy = copy(sampleTable)

    subjIDPhenList =[]
    phenTitleList = []

    firstTime = True
    for row in samplePhenotypeTable.sheet_by_index(0)._cell_values:
        if firstTime:
            phenTitleList = row
            firstTime = False
        subjIDPhen = row [0]
        subjIDPhenList.append(subjIDPhen)

    sampleTableCopy.get_sheet(0).write(0,61,phenTitleList[0])
    sampleTableCopy.get_sheet(0).write(0,62,phenTitleList[1])
    sampleTableCopy.get_sheet(0).write(0,63,phenTitleList[2])
    sampleTableCopy.get_sheet(0).write(0,64,phenTitleList[3])

    j=0
    for row in sampleTable.sheet_by_index(0)._cell_values:
        print j
        i=0
        subjIDGen = ''
        for c in row[0]:
            subjIDGen += c
            if c == '-':
                i+=1
                if i == 2:
                    subjIDGen = subjIDGen[:-1]
                    try:
                        rowContent = samplePhenotypeTable.sheet_by_index(0)._cell_values[subjIDPhenList.index(subjIDGen)]
                    except ValueError:
                        print 'SUBJID : ', subjIDGen, ' not find in the phenotype table'
                    sampleTableCopy.get_sheet(0).write(j,61,rowContent[0])
                    sampleTableCopy.get_sheet(0).write(j,62,rowContent[1])
                    sampleTableCopy.get_sheet(0).write(j,63,rowContent[2])
                    sampleTableCopy.get_sheet(0).write(j,64,rowContent[3])
                    break
        j+=1

    sampleTableCopy.save(sampleTableFileName)
    return

def writeGeneTable(nameList, descriptionList):
    '''write table (excel) which contain Gene informations'''
    geneTable = open_workbook(geneTableFileName)
    geneTableCopy = copy(geneTable)

    i=0
    while (i < len(nameList)):
        print i
        geneTableCopy.get_sheet(0).write(i+1, 0, nameList[i])
        geneTableCopy.get_sheet(0).write(i+1, 1, descriptionList[i])
        i+=1
    geneTableCopy.save(geneTableFileNameTest)

    return

def updateGeneTable(geneList, chrList, coordList):
    '''add information in the Gene table'''
    geneTable = open_workbook(geneTableFileName)
    geneTableCopy = copy(geneTable)

    rowIndex = 0
    for row in geneTable.sheet_by_index(0)._cell_values:
        print rowIndex
        try:
            listIndex = geneList.index(row[0])
            geneTableCopy.get_sheet(0).write(rowIndex, 2, chrList[listIndex])
            geneTableCopy.get_sheet(0).write(rowIndex, 3, coordList[listIndex])
        except ValueError:
            print 'GeneSymbol: ', row[0], ' not find in the gene table'

        rowIndex +=1

    geneTableCopy.save(geneTableFileNameTest)

    return

def writeTranscriptTable(transcriptList, geneList):
    transcriptTable = openpyxl.Workbook()
    transcriptTableSheet = transcriptTable.get_active_sheet()

    rowIndex = 0
    while(rowIndex < len(transcriptList)):
        print 'rowIndex: ', rowIndex
        transcriptTableSheet.cell(row = rowIndex, column = 0).value = transcriptList[rowIndex]
        transcriptTableSheet.cell(row = rowIndex, column = 1).value = geneList[rowIndex]

        rowIndex +=1

    transcriptTable.save(transcriptTableFileName)
    return

def writeRPKMGeneTable():

    csvFileReader = open(geneRPKMFileNameTail, 'rb')
    geneRKMTable = csv.reader(csvFileReader, delimiter = '	')

    csvFileWriter = open(geneRPKMTableFileName, 'wb')
    outputWriter = csv.writer(csvFileWriter, quotechar = '"', quoting = csv.QUOTE_MINIMAL)

    outputWriter.writerow(["RPKMID", "GeneID", "SAMPID", "RPKMValue"])

    rowIndex = 0
    newRowIndex = 0

    sampList = []

    for row in geneRKMTable:
        print 'rowIndex: ', rowIndex
        colIndex = 0
        for element in row[2:]:
            if (rowIndex==0):
                sampList.append(element)
            else:
                outputWriter.writerow([newRowIndex, row[0], sampList[colIndex], element ])
                newRowIndex +=1
            colIndex +=1
        rowIndex+=1
    return

def writeRPKMTranscriptTable():

    csvFileReader = open(transcriptRPKMFileName, 'rb')
    geneRKMTable = csv.reader(csvFileReader, delimiter = '	')

    csvFileWriter = open(transcriptRPKMTableFileNameExterne, 'wb')
    outputWriter = csv.writer(csvFileWriter, quotechar = '"', quoting = csv.QUOTE_MINIMAL)

    outputWriter.writerow(["RPKMID", "TranscriptID", "SAMPID", "RPKMValue"])

    rowIndex = 0
    newRowIndex = 0

    sampList = []

    for row in geneRKMTable:
        print 'rowIndex: ', rowIndex
        colIndex = 0
        for element in row[4:]:
            if (rowIndex==0):
                sampList.append(element)
            else:
                outputWriter.writerow([newRowIndex, row[0], sampList[colIndex], element ])
                newRowIndex +=1
            colIndex +=1
        rowIndex+=1
    return

def removeLast(string):
    return string[:-1]


def main():
    '''main function'''
    startTime = time.time()


    #writeSampleTable()
    # writeRPKMTranscriptTable()

    endTime = time.time()-startTime
    print 'endTime : ', endTime
    return

def processARNTable():
    writeSampleTable()
    return

def processGeneTable():
    #make table
    sampleList, geneNameList, geneDescriptionList = extractGene(geneRPKMFileName)
    geneNameList = geneNameList[1:]
    cleanList(sampleList, 'sample')
    cleanList(sampleList, 'sample')
    writeGeneTable(geneNameList,geneDescriptionList)

    print 'sampleList: ', sampleList, ' length: ', len(sampleList)
    print 'nameList: ', geneNameList, ' length: ', len(geneNameList)
    print 'descriptionList', geneDescriptionList, 'length: ', len(geneDescriptionList)

    #update table

    transcript, gene, chr, coord = extractTranscript(transcriptRPKMFileNameTest)


    gene = gene[1:]
    chr = chr[1:]
    coord = coord[1:]

    gene = map(removeLast, gene)
    chr = map(removeLast, chr) # a faire plus tard
    coord = map(removeLast, coord)

    # print 'gene : ', gene
    # print 'chr : ', chr
    # print 'coord : ', coord

    updateGeneTable(gene, chr, coord)

    return

def processTranscriptTable():
    transcript, gene, chr, coord = extractTranscript(transcriptRPKMFileName)

    gene = map(removeLast, gene)
    transcript = map(removeLast, transcript)

    writeTranscriptTable(transcript, gene)
    return

def processRPKMGeneTable():
    writeRPKMGeneTable()
    return

main()